import pandas as pd
import os
import subprocess
import platform
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import sys

# 添加项目根目录到Python路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    from utils.logger import logger
except ImportError:
    import logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(__name__)

class ExcelExporter:
    """Excel导出类，负责将数据导出为格式化的Excel文件"""
    
    def __init__(self):
        # 定义样式
        self.header_font = Font(name='微软雅黑', size=12, bold=True, color='000000')
        self.content_font = Font(name='微软雅黑', size=11, color='000000')
        
        self.border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 高亮填充
        self.profit_rate_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        self.modified_price_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        self.unmatched_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # 列宽设置 - 统一设置为15字符
        self.profit_table_widths = {
            '配置': 25,
            '速别': 15,
            '简称': 45,
            '价格': 15,
            '成本': 15,
            '快递': 15,
            '毛利润': 15,
            '毛利率': 15
        }
        
        # 原始数据列宽设置 - 增加宽度
        self.original_data_widths = {
            '货品ID': 25,
            '规格ID': 25,
            '分类': 18,
            '简称': 35,
            '价格': 15,
            '成本': 15,
            '毛利': 15,
            '毛利率': 15
        }

    def export_profit_table(self, file_path: str, profit_table: pd.DataFrame, original_data: pd.DataFrame = None):
        """导出毛利表到Excel文件"""
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 调整列顺序，将简称列移到最右侧
                reordered_table = self._reorder_profit_table_columns(profit_table)
                
                # 写入毛利表
                reordered_table.to_excel(writer, sheet_name='毛利表', index=False)
                
                # 美化毛利表
                self._format_profit_table(writer, reordered_table)
                
                # 如果有原始数据，也写入
                if original_data is not None:
                    original_data.to_excel(writer, sheet_name='原始数据', index=False)
                
                logger.info(f"毛利表已导出到: {file_path}")
                
                # 导出完成提示和打开选项
                self._show_export_completion(file_path, "毛利表")
                
        except Exception as e:
            logger.error(f"导出毛利表失败: {e}")
            raise

    def _reorder_profit_table_columns(self, profit_table: pd.DataFrame) -> pd.DataFrame:
        """重新排列毛利表列顺序，将简称列移到最右侧"""
        # 定义期望的列顺序（简称放在最后）
        desired_order = ['配置', '速别', '价格', '成本', '快递', '毛利润', '毛利率', '简称']
        
        # 获取实际存在的列
        existing_cols = [col for col in desired_order if col in profit_table.columns]
        
        # 添加任何不在期望顺序中的列
        remaining_cols = [col for col in profit_table.columns if col not in existing_cols]
        final_order = existing_cols + remaining_cols
        
        return profit_table[final_order]

    def export_original_data(self, file_path: str, data: pd.DataFrame):
        """导出原始数据到Excel文件"""
        try:
            # 准备导出数据：移除"未匹配"列，确保"新毛利率"列存在
            export_data = data.copy()
            
            # 移除"未匹配"列（如果存在）
            if '未匹配' in export_data.columns:
                export_data = export_data.drop(columns=['未匹配'])
            
            # 如果没有"新毛利率"列，添加一个空列
            if '新毛利率' not in export_data.columns:
                export_data['新毛利率'] = "无法计算"
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 导出数据
                export_data.to_excel(writer, sheet_name='修改后原始数据', index=False)
                
                # 格式化原始数据表
                self._format_original_data_table(writer, export_data, data)
                
                logger.info(f"原始数据已导出到: {file_path}")
                
                # 导出完成提示和打开选项
                self._show_export_completion(file_path, "原始数据")
                
        except Exception as e:
            logger.error(f"导出原始数据失败: {e}")
            raise

    def _format_profit_table(self, writer, profit_table: pd.DataFrame):
        """格式化毛利表"""
        workbook = writer.book
        worksheet = writer.sheets['毛利表']
        
        # 在表头上方插入一行
        worksheet.insert_rows(1)
        
        # 设置列宽
        for col_num, column_title in enumerate(profit_table.columns, 1):
            col_letter = get_column_letter(col_num)
            if column_title in self.profit_table_widths:
                worksheet.column_dimensions[col_letter].width = self.profit_table_widths[column_title]
        
        # 找到各列的索引
        config_col = profit_table.columns.get_loc('配置') + 1 if '配置' in profit_table.columns else None
        speed_col = profit_table.columns.get_loc('速别') + 1 if '速别' in profit_table.columns else None
        name_col = profit_table.columns.get_loc('简称') + 1 if '简称' in profit_table.columns else None
        price_col = profit_table.columns.get_loc('价格') + 1 if '价格' in profit_table.columns else None
        cost_col = profit_table.columns.get_loc('成本') + 1 if '成本' in profit_table.columns else None
        express_col = profit_table.columns.get_loc('快递') + 1 if '快递' in profit_table.columns else None
        profit_amount_col = profit_table.columns.get_loc('毛利润') + 1 if '毛利润' in profit_table.columns else None
        profit_rate_col = profit_table.columns.get_loc('毛利率') + 1 if '毛利率' in profit_table.columns else None
        
        # 合并配置、速别、价格列的第一行和第二行单元格，并设置表头名称
        merge_cols_info = [
            (config_col, '配置'),
            (speed_col, '速别'), 
            (price_col, '价格')
        ]
        for col_idx, col_name in merge_cols_info:
            if col_idx is not None:
                try:
                    worksheet.merge_cells(start_row=1, start_column=col_idx, 
                                        end_row=2, end_column=col_idx)
                    # 设置合并后单元格的表头名称和样式
                    header_cell = worksheet.cell(row=1, column=col_idx, value=col_name)
                    header_cell.alignment = self.center_alignment
                    header_cell.font = self.header_font
                    header_cell.border = self.border_thin
                    logger.info(f"成功合并列 {col_name} (索引: {col_idx})")
                except Exception as e:
                    logger.error(f"合并列 {col_name} 失败: {e}")
        
        # 合并费用分组标题（成本、快递）
        if cost_col and express_col:
            worksheet.merge_cells(start_row=1, start_column=cost_col, 
                                end_row=1, end_column=express_col)
            fee_cell = worksheet.cell(row=1, column=cost_col, value='费用')
            fee_cell.alignment = self.center_alignment
            fee_cell.font = self.header_font
            fee_cell.border = self.border_thin
        
        # 合并毛利分组标题（毛利润、毛利率）
        if profit_amount_col and profit_rate_col:
            worksheet.merge_cells(start_row=1, start_column=profit_amount_col, 
                                end_row=1, end_column=profit_rate_col)
            profit_cell = worksheet.cell(row=1, column=profit_amount_col, value='毛利')
            profit_cell.alignment = self.center_alignment
            profit_cell.font = self.header_font
            profit_cell.border = self.border_thin
        
        # 简称列现在在最右侧，需要跨两行合并
        if name_col:
            worksheet.merge_cells(start_row=1, start_column=name_col, 
                                end_row=2, end_column=name_col)
            name_header_cell = worksheet.cell(row=1, column=name_col, value='简称')
            name_header_cell.alignment = self.center_alignment
            name_header_cell.font = self.header_font
            name_header_cell.border = self.border_thin
        
        # 应用表头样式（第2行）
        for col_num, column_title in enumerate(profit_table.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border_thin
            
            # 毛利率列高亮
            if column_title == '毛利率':
                cell.fill = self.profit_rate_fill
        
        # 应用数据行样式并设置公式（数据从第3行开始）
        for row_num in range(3, len(profit_table) + 3):
            # 设置行高
            worksheet.row_dimensions[row_num].height = 20
            for col_num in range(1, len(profit_table.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = self.content_font
                cell.alignment = self.center_alignment
                cell.border = self.border_thin
                
                column_name = profit_table.columns[col_num-1]
                
                # 设置毛利润公式：价格 - 成本 - 快递
                if column_name == '毛利润':
                    price_col_letter = get_column_letter(price_col) if price_col else None
                    cost_col_letter = get_column_letter(cost_col) if cost_col else None
                    express_col_letter = get_column_letter(express_col) if express_col else None
                    
                    if price_col_letter and cost_col_letter and express_col_letter:
                        formula = f"={price_col_letter}{row_num}-{cost_col_letter}{row_num}-{express_col_letter}{row_num}"
                        cell.value = formula
                
                # 设置毛利率公式：毛利润 / 价格 * 100%
                elif column_name == '毛利率':
                    price_col_letter = get_column_letter(price_col) if price_col else None
                    profit_col_letter = get_column_letter(profit_amount_col) if profit_amount_col else None
                    
                    if price_col_letter and profit_col_letter:
                        formula = f"=IF({price_col_letter}{row_num}=0,0,{profit_col_letter}{row_num}/{price_col_letter}{row_num})"
                        cell.value = formula
                        cell.number_format = '0.00%'  # 设置为百分比格式
                    
                    cell.fill = self.profit_rate_fill
                

        
        # 合并相同分类的单元格（数据从第3行开始）
        self._merge_category_cells(worksheet, profit_table, start_row=3)
        
        # 不再自动调整列宽，使用预设的22字符宽度

    def _merge_category_cells(self, worksheet, profit_table: pd.DataFrame, start_row=3):
        """合并相同配置名的单元格"""
        if '配置' not in profit_table.columns:
            return
            
        category_col_idx = profit_table.columns.get_loc('配置') + 1  # Excel列从1开始
        
        current_config = None
        merge_start_row = start_row  # 数据开始行
        merge_ranges = []
        
        for row_idx in range(start_row, len(profit_table) + start_row):
            category_value = worksheet.cell(row=row_idx, column=category_col_idx).value
            
            if category_value != current_config:
                if current_config is not None and (row_idx - merge_start_row) > 1:
                    # 需要合并的行范围
                    merge_ranges.append((merge_start_row, row_idx - 1))
                current_config = category_value
                merge_start_row = row_idx
        
        # 处理最后一组
        if current_config is not None and (len(profit_table) + start_row - merge_start_row) > 1:
            merge_ranges.append((merge_start_row, len(profit_table) + start_row - 1))
        
        # 执行合并
        for start, end in merge_ranges:
            if end > start:  # 只有多行才需要合并
                worksheet.merge_cells(start_row=start, start_column=category_col_idx, 
                                    end_row=end, end_column=category_col_idx)
                # 设置合并后单元格的垂直居中
                cell = worksheet.cell(row=start, column=category_col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def _auto_adjust_column_widths(self, worksheet, profit_table: pd.DataFrame, header_row=2):
        """自动调整列宽"""
        for col_num, column_title in enumerate(profit_table.columns, 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            
            # 获取表头宽度（现在表头在第2行）
            header_cell = worksheet.cell(row=header_row, column=col_num)
            header_length = len(str(header_cell.value)) if header_cell.value else 0
            max_length = max(max_length, header_length)
            
            # 获取数据行最大宽度（数据从第3行开始）
            for row_num in range(3, len(profit_table) + 3):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell_value = str(cell.value) if cell.value else ""
                cell_length = len(cell_value)
                max_length = max(max_length, cell_length)
            
            # 设置列宽，留一些边距
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            worksheet.column_dimensions[col_letter].width = adjusted_width

    def _format_original_data_table(self, writer, export_data: pd.DataFrame, original_data: pd.DataFrame = None):
        """格式化原始数据表"""
        workbook = writer.book
        worksheet = writer.sheets['修改后原始数据']
        
        # 更新列宽设置，添加新毛利率列
        updated_widths = self.original_data_widths.copy()
        updated_widths['新毛利率'] = 15
        updated_widths['修改后价格'] = 15
        
        # 设置列宽
        for col_num, column_title in enumerate(export_data.columns, 1):
            col_letter = get_column_letter(col_num)
            if column_title in updated_widths:
                worksheet.column_dimensions[col_letter].width = updated_widths[column_title]
            else:
                worksheet.column_dimensions[col_letter].width = 15
        
        # 应用表头样式
        for col_num, column_title in enumerate(export_data.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border_thin
        
        # 应用数据行样式
        for row_num in range(2, len(export_data) + 2):
            for col_num in range(1, len(export_data.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = self.content_font
                cell.alignment = self.center_alignment
                cell.border = self.border_thin
                
                column_name = export_data.columns[col_num-1]
                data_row_idx = row_num - 2  # 数据起始于第2行
                
                # ID列设置为文本格式，防止精度丢失
                if 'ID' in column_name or 'id' in column_name or '货品ID' in column_name or '规格ID' in column_name:
                    cell.number_format = '@'  # 文本格式
                
                # 修改后价格列通过背景颜色显示匹配状态
                if column_name == '修改后价格':
                    is_unmatched = False
                    try:
                        # 从原始数据中获取匹配状态
                        if original_data is not None and '未匹配' in original_data.columns and 0 <= data_row_idx < len(original_data):
                            unmatched_col_idx = original_data.columns.get_loc('未匹配')
                            is_unmatched = bool(original_data.iloc[data_row_idx, unmatched_col_idx])
                        else:
                            # 如果没有原始数据，通过价格值判断
                            cell_value = export_data.iloc[data_row_idx, col_num-1]
                            is_unmatched = pd.isna(cell_value) or cell_value is None
                    except:
                        is_unmatched = True
                    
                    cell.fill = self.unmatched_fill if is_unmatched else self.modified_price_fill
                
                # 新毛利率列样式
                elif column_name == '新毛利率':
                    # 为新毛利率列设置百分比样式的背景色
                    profit_rate_light_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')  # 淡蓝色背景
                    cell.fill = profit_rate_light_fill

    def _show_export_completion(self, file_path: str, file_type: str):
        """显示导出完成提示并提供打开文件选项"""
        import tkinter as tk
        from tkinter import messagebox
        
        print(f"✅ {file_type}导出完成！")
        print(f"📁 文件位置: {file_path}")
        
        # 显示消息框询问是否打开文件
        message = f"✅ {file_type}导出完成！📁 文件位置:{file_path}是否要打开文件？"

        result = messagebox.askyesno("导出完成", message)
        
        if result:
            self._open_file(file_path)

    def _open_file(self, file_path: str):
        """跨平台打开文件"""
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(file_path)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", file_path])
            else:  # Linux
                subprocess.run(["xdg-open", file_path])
            print(f"✅ 已打开文件: {file_path}")
        except Exception as e:
            print(f"❌ 打开文件失败: {e}")
            print(f"请手动打开文件: {file_path}")